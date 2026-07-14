import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# Style definitions
# ============================================================
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(name='Microsoft YaHei', size=10, bold=True)
title_font = Font(name='Microsoft YaHei', size=12, bold=True)
normal_font = Font(name='Microsoft YaHei', size=9)
bold_font = Font(name='Microsoft YaHei', size=9, bold=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

title_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
header_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
example_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
synthesis_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
ctrl_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
blank_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
group_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

def apply_style(ws, row, col, value, font=normal_font, fill=None, align=center_align):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = align
    cell.border = thin_border
    if fill:
        cell.fill = fill
    return cell

def apply_merge(ws, r1, c1, r2, c2, value, font=normal_font, fill=None, align=center_align):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = apply_style(ws, r1, c1, value, font, fill, align)
    for rr in range(r1, r2+1):
        for cc in range(c1, c2+1):
            ws.cell(row=rr, column=cc).border = thin_border
            if fill:
                ws.cell(row=rr, column=cc).fill = fill
    return cell

DASH = '-'
DEMO_SET = {'addi','ori','slli','lui','lw','beq','bne','jal'}

# ============================================================
# SHEET 1: 数据通路表(含控制信号)
# ============================================================
ws1 = wb.active
ws1.title = "数据通路表(含控制信号)"

col_w1 = {1:5, 2:10, 3:12, 4:12, 5:12, 6:12, 7:14, 8:14, 9:14, 10:22, 11:22, 12:15, 13:15, 14:15, 15:15, 16:18, 17:14}
for c, w in col_w1.items():
    ws1.column_dimensions[get_column_letter(c)].width = w

# Title
r = 2
apply_merge(ws1, r, 2, r, 17, '数据通路表 - miniRV (37条指令)', title_font, title_fill)

# Category row
r = 3
apply_merge(ws1, r, 2, r, 2, '功能单元', header_font, header_fill)
apply_merge(ws1, r, 3, r, 6, '取指单元', header_font, header_fill)
apply_merge(ws1, r, 7, r, 10, '译码单元', header_font, header_fill)
apply_merge(ws1, r, 11, r, 11, '扩展单元', header_font, header_fill)
apply_merge(ws1, r, 12, r, 13, '执行单元', header_font, header_fill)
apply_merge(ws1, r, 14, r, 15, '访存请求', header_font, header_fill)
apply_merge(ws1, r, 16, r, 17, '访存扩展', header_font, header_fill)

# Component row
r = 4
apply_style(ws1, r, 2, '部件', header_font, header_fill)
apply_style(ws1, r, 3, 'PC', header_font, header_fill)
apply_merge(ws1, r, 4, r, 6, 'NPC', header_font, header_fill)
apply_merge(ws1, r, 7, r, 10, 'RF', header_font, header_fill)
apply_style(ws1, r, 11, 'SEXT', header_font, header_fill)
apply_merge(ws1, r, 12, r, 13, 'ALU', header_font, header_fill)
apply_merge(ws1, r, 14, r, 15, 'MREQ', header_font, header_fill)
apply_merge(ws1, r, 16, r, 17, 'MEXT', header_font, header_fill)

# Signal row
r = 5
apply_style(ws1, r, 2, '模块信号', header_font, header_fill)
sig_h = {3:'npc', 4:'pc', 5:'offset', 6:'br', 7:'rR1', 8:'rR2', 9:'wR', 10:'wD',
         11:'imm', 12:'A', 13:'B', 14:'ram_addr', 15:'ram_wdata', 16:'din', 17:'byte_offs'}
for c, v in sig_h.items():
    apply_style(ws1, r, c, v, header_font, header_fill)

# ============================================================
# Instruction data
# Each tuple: (instruction, npc, pc, offset, br, rR1, rR2, wR, wD, imm, A, B, ram_addr, ram_wdata, din, byte_offs)
# ============================================================
inst_data = [
    # === Group label ===
    ('__GROUP__', '8条示例指令 (模板工程已实现)'),
    ('addi',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', DASH, DASH, DASH, DASH),
    ('ori',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', DASH, DASH, DASH, DASH),
    ('slli',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', DASH, DASH, DASH, DASH),
    ('lui',   'NPC.npc', 'PC.pc', DASH, DASH, DASH, DASH, 'IN.inst[11:7]', 'SEXT.ext', 'IN.inst[31:12]', DASH, DASH, DASH, DASH, DASH, DASH),
    ('lw',    'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'MEXT.ext', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', 'ALU.C', DASH, 'IN.daccess_rdata', 'ALU.C[1:0]'),
    ('beq',   'NPC.npc', 'PC.pc', 'SEXT.ext', 'ALU.br', 'IN.inst[19:15]', 'IN.inst[24:20]', DASH, DASH, 'IN.inst[31|7|30:25|11:8]', 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('bne',   'NPC.npc', 'PC.pc', 'SEXT.ext', 'ALU.br', 'IN.inst[19:15]', 'IN.inst[24:20]', DASH, DASH, 'IN.inst[31|7|30:25|11:8]', 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('jal',   'NPC.npc', 'PC.pc', 'SEXT.ext', DASH, DASH, DASH, 'IN.inst[11:7]', 'NPC.pc4', 'IN.inst[31|19:12|20|30:21]', DASH, DASH, DASH, DASH, DASH, DASH),

    # === R-type arithmetic ===
    ('__GROUP__', 'R型算逻指令 (10条)'),
    ('add',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('sub',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('and',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('or',    'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('xor',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('sll',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('srl',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('sra',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('slt',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('sltu',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),

    # === R-type multiply/divide ===
    ('__GROUP__', 'R型乘除指令 (7条, M扩展)'),
    ('mul',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('mulh',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('mulhu', 'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('div',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('divu',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('rem',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('remu',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]', 'ALU.C', DASH, 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),

    # === I-type arithmetic ===
    ('__GROUP__', 'I型算逻指令 (6条, 除addi/ori/slli外)'),
    ('andi',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', DASH, DASH, DASH, DASH),
    ('xori',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', DASH, DASH, DASH, DASH),
    ('srli',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', DASH, DASH, DASH, DASH),
    ('srai',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', DASH, DASH, DASH, DASH),
    ('slti',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', DASH, DASH, DASH, DASH),
    ('sltiu', 'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', DASH, DASH, DASH, DASH),

    # === I-type load ===
    ('__GROUP__', 'I型访存指令 (5条, 除lw外) + JALR'),
    ('lb',    'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'MEXT.ext', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', 'ALU.C', DASH, 'IN.daccess_rdata', 'ALU.C[1:0]'),
    ('lbu',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'MEXT.ext', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', 'ALU.C', DASH, 'IN.daccess_rdata', 'ALU.C[1:0]'),
    ('lh',    'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'MEXT.ext', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', 'ALU.C', DASH, 'IN.daccess_rdata', 'ALU.C[1:0]'),
    ('lhu',   'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'MEXT.ext', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', 'ALU.C', DASH, 'IN.daccess_rdata', 'ALU.C[1:0]'),
    ('jalr',  'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', DASH, 'IN.inst[11:7]', 'NPC.pc4', 'IN.inst[31:20]', 'RF.rD1', 'SEXT.ext', 'ALU.C', DASH, DASH, DASH),

    # === S-type store ===
    ('__GROUP__', 'S型存储指令 (3条)'),
    ('sb',    'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', DASH, DASH, 'IN.inst[31:25|11:7]', 'RF.rD1', 'SEXT.ext', 'ALU.C', 'RF.rD2', DASH, DASH),
    ('sh',    'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', DASH, DASH, 'IN.inst[31:25|11:7]', 'RF.rD1', 'SEXT.ext', 'ALU.C', 'RF.rD2', DASH, DASH),
    ('sw',    'NPC.npc', 'PC.pc', DASH, DASH, 'IN.inst[19:15]', 'IN.inst[24:20]', DASH, DASH, 'IN.inst[31:25|11:7]', 'RF.rD1', 'SEXT.ext', 'ALU.C', 'RF.rD2', DASH, DASH),

    # === B-type branch (additional 4) ===
    ('__GROUP__', 'B型分支指令 (4条, 除beq/bne外)'),
    ('blt',   'NPC.npc', 'PC.pc', 'SEXT.ext', 'ALU.br', 'IN.inst[19:15]', 'IN.inst[24:20]', DASH, DASH, 'IN.inst[31|7|30:25|11:8]', 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('bltu',  'NPC.npc', 'PC.pc', 'SEXT.ext', 'ALU.br', 'IN.inst[19:15]', 'IN.inst[24:20]', DASH, DASH, 'IN.inst[31|7|30:25|11:8]', 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('bge',   'NPC.npc', 'PC.pc', 'SEXT.ext', 'ALU.br', 'IN.inst[19:15]', 'IN.inst[24:20]', DASH, DASH, 'IN.inst[31|7|30:25|11:8]', 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),
    ('bgeu',  'NPC.npc', 'PC.pc', 'SEXT.ext', 'ALU.br', 'IN.inst[19:15]', 'IN.inst[24:20]', DASH, DASH, 'IN.inst[31|7|30:25|11:8]', 'RF.rD1', 'RF.rD2', DASH, DASH, DASH, DASH),

    # === U-type ===
    ('__GROUP__', 'U型指令 (1条, 除lui外)'),
    ('auipc', 'NPC.npc', 'PC.pc', DASH, DASH, DASH, DASH, 'IN.inst[11:7]', 'ALU.C', 'IN.inst[31:12]', 'PC.pc', 'SEXT.ext', DASH, DASH, DASH, DASH),
]

# Write instruction rows
r = 6
for item in inst_data:
    if item[0] == '__GROUP__':
        # Group label row
        apply_merge(ws1, r, 2, r, 17, item[1], bold_font, group_fill, left_align)
        r += 1
        continue

    inst_name = item[0]
    values = item[1:]
    fill = example_fill if inst_name in DEMO_SET else blank_fill

    apply_style(ws1, r, 2, inst_name, normal_font, fill, left_align)
    for i, v in enumerate(values):
        apply_style(ws1, r, 3+i, v, normal_font, fill, center_align)
    r += 1

# --- 综合 row ---
r_syn = r
apply_style(ws1, r, 2, '综合', bold_font, synthesis_fill, left_align)
syn_vals = [
    'NPC.npc', 'PC.pc', 'SEXT.ext', 'ALU.br',
    'IN.inst[19:15]', 'IN.inst[24:20]', 'IN.inst[11:7]',
    'ALU.C\nSEXT.ext\nMEXT.ext\nNPC.pc4',
    'IN.inst[31:7]',
    'RF.rD1\nPC.pc',
    'RF.rD2\nSEXT.ext',
    'ALU.C', 'RF.rD2', 'IN.daccess_rdata', 'ALU.C[1:0]'
]
for i, v in enumerate(syn_vals):
    apply_style(ws1, r, 3+i, v, normal_font, synthesis_fill, center_align)
r += 1

# --- 操作选择信号 row ---
apply_style(ws1, r, 2, '操作选择信号', bold_font, ctrl_fill, left_align)
op_sel = {4:'npc_op', 9:'rf_we', 11:'sext_op', 12:'alu_op', 14:'ram_rop', 15:'ram_wop', 16:'ram_rop'}
for c in range(3, 18):
    v = op_sel.get(c, DASH)
    apply_style(ws1, r, c, v, normal_font, ctrl_fill, center_align)
r += 1

# --- 多路选择信号 row ---
apply_style(ws1, r, 2, '多路选择信号', bold_font, ctrl_fill, left_align)
mux_sel = {10:'rf_wsel', 12:'alua_sel', 13:'alub_sel'}
for c in range(3, 18):
    v = mux_sel.get(c, DASH)
    apply_style(ws1, r, c, v, normal_font, ctrl_fill, center_align)

for row in range(3, r+1):
    ws1.row_dimensions[row].height = 22
ws1.row_dimensions[r_syn].height = 55

# ============================================================
# SHEET 2: 控制信号取值表
# ============================================================
ws2 = wb.create_sheet("控制信号取值表")

col_w2 = {1:3, 2:8, 3:10, 4:10, 5:10, 6:10, 7:7, 8:10, 9:10, 10:10, 11:10, 12:10, 13:10, 14:12, 15:12}
for c, w in col_w2.items():
    ws2.column_dimensions[get_column_letter(c)].width = w

r = 2
apply_merge(ws2, r, 2, r, 15, '控制信号取值表 - miniRV (37条指令)', title_font, title_fill)

# Headers
r = 3
apply_merge(ws2, r, 2, r+1, 2, '指令', header_font, header_fill)
apply_merge(ws2, r, 3, r+1, 3, 'opcode', header_font, header_fill)
apply_merge(ws2, r, 4, r+1, 4, 'funct3', header_font, header_fill)
apply_merge(ws2, r, 5, r+1, 5, 'funct7', header_font, header_fill)
apply_merge(ws2, r, 6, r, 15, '控制信号', header_font, header_fill)

r = 4
ctrl_h = {6:'npc_op', 7:'rf_we', 8:'rf_wsel', 9:'sext_op', 10:'alu_op',
          11:'alua_sel', 12:'alub_sel', 13:'npc_op\n(B-type)', 14:'ram_rop', 15:'ram_wop'}
for c, v in ctrl_h.items():
    apply_style(ws2, r, c, v, header_font, header_fill)

# Control data
ctrl_data = [
    # __GROUP__, instr, opcode, funct3, funct7, npc_op, rf_we, rf_wsel, sext_op, alu_op, alua_sel, alub_sel, branch_npc, ram_rop, ram_wop
    ('__GROUP__', '8条示例指令'),
    ('addi',  '0010011', '000', '-',    'PC4', '1', 'WB_ALU', 'I_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', '-',    '0',       '0'),
    ('ori',   '0010011', '110', '-',    'PC4', '1', 'WB_ALU', 'I_TYPE', 'OR',  'SEL_RS1', 'SEL_EXT', '-',    '0',       '0'),
    ('slli',  '0010011', '001', '0000000','PC4','1', 'WB_ALU', 'I_TYPE', 'SLL', 'SEL_RS1', 'SEL_EXT', '-',    '0',       '0'),
    ('lui',   '0110111', '-',   '-',    'PC4', '1', 'WB_EXT', 'U_TYPE', '-',   '-',       '-',       '-',    '0',       '0'),
    ('lw',    '0000011', '010', '-',    'PC4', '1', 'WB_RAM', 'I_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', '-',    'WORD_EXT','0'),
    ('beq',   '1100011', '000', '-',    'PC4', '0', '-',      'B_TYPE', 'EQ',  'SEL_RS1', 'SEL_RS2', 'BRANCH','0',      '0'),
    ('bne',   '1100011', '001', '-',    'PC4', '0', '-',      'B_TYPE', 'NE',  'SEL_RS1', 'SEL_RS2', 'BRANCH','0',      '0'),
    ('jal',   '1101111', '-',   '-',    'PC4', '1', 'WB_PC4', 'J_TYPE', '-',   '-',       '-',       'JUMP',  '0',       '0'),

    ('__GROUP__', 'R型算逻 (10条)'),
    ('add',   '0110011', '000', '0000000','PC4','1', 'WB_ALU', '-',      'ADD', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('sub',   '0110011', '000', '0100000','PC4','1', 'WB_ALU', '-',      'SUB', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('and',   '0110011', '111', '0000000','PC4','1', 'WB_ALU', '-',      'AND', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('or',    '0110011', '110', '0000000','PC4','1', 'WB_ALU', '-',      'OR',  'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('xor',   '0110011', '100', '0000000','PC4','1', 'WB_ALU', '-',      'XOR', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('sll',   '0110011', '001', '0000000','PC4','1', 'WB_ALU', '-',      'SLL', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('srl',   '0110011', '101', '0000000','PC4','1', 'WB_ALU', '-',      'SRL', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('sra',   '0110011', '101', '0100000','PC4','1', 'WB_ALU', '-',      'SRA', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('slt',   '0110011', '010', '0000000','PC4','1', 'WB_ALU', '-',      'SLT', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('sltu',  '0110011', '011', '0000000','PC4','1', 'WB_ALU', '-',      'SLTU','SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),

    ('__GROUP__', 'R型乘除 (7条, M扩展)'),
    ('mul',   '0110011', '000', '0000001','PC4','1', 'WB_ALU', '-',      'MUL', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('mulh',  '0110011', '001', '0000001','PC4','1', 'WB_ALU', '-',      'MULH','SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('mulhu', '0110011', '011', '0000001','PC4','1', 'WB_ALU', '-',      'MULHU','SEL_RS1','SEL_RS2','-',   '0',       '0'),
    ('div',   '0110011', '100', '0000001','PC4','1', 'WB_ALU', '-',      'DIV', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('divu',  '0110011', '101', '0000001','PC4','1', 'WB_ALU', '-',      'DIVU','SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('rem',   '0110011', '110', '0000001','PC4','1', 'WB_ALU', '-',      'REM', 'SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),
    ('remu',  '0110011', '111', '0000001','PC4','1', 'WB_ALU', '-',      'REMU','SEL_RS1', 'SEL_RS2', '-',    '0',       '0'),

    ('__GROUP__', 'I型算逻 (6条)'),
    ('andi',  '0010011', '111', '-',     'PC4', '1', 'WB_ALU', 'I_TYPE', 'AND', 'SEL_RS1', 'SEL_EXT', '-',    '0',       '0'),
    ('xori',  '0010011', '100', '-',     'PC4', '1', 'WB_ALU', 'I_TYPE', 'XOR', 'SEL_RS1', 'SEL_EXT', '-',    '0',       '0'),
    ('srli',  '0010011', '101', '0000000','PC4','1', 'WB_ALU', 'I_TYPE', 'SRL', 'SEL_RS1', 'SEL_EXT', '-',    '0',       '0'),
    ('srai',  '0010011', '101', '0100000','PC4','1', 'WB_ALU', 'I_TYPE', 'SRA', 'SEL_RS1', 'SEL_EXT', '-',    '0',       '0'),
    ('slti',  '0010011', '010', '-',     'PC4', '1', 'WB_ALU', 'I_TYPE', 'SLT', 'SEL_RS1', 'SEL_EXT', '-',    '0',       '0'),
    ('sltiu', '0010011', '011', '-',     'PC4', '1', 'WB_ALU', 'I_TYPE', 'SLTU','SEL_RS1', 'SEL_EXT', '-',    '0',       '0'),

    ('__GROUP__', 'I型访存 + JALR (6条)'),
    ('lb',    '0000011', '000', '-',     'PC4', '1', 'WB_RAM', 'I_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', '-',    'BYTE_EXT', '0'),
    ('lbu',   '0000011', '100', '-',     'PC4', '1', 'WB_RAM', 'I_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', '-',    'BYTEU_EXT','0'),
    ('lh',    '0000011', '001', '-',     'PC4', '1', 'WB_RAM', 'I_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', '-',    'HALF_EXT', '0'),
    ('lhu',   '0000011', '101', '-',     'PC4', '1', 'WB_RAM', 'I_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', '-',    'HALFU_EXT','0'),
    ('jalr',  '1100111', '000', '-',     'PC4', '1', 'WB_PC4', 'I_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', 'JUMP',  '0',       '0'),

    ('__GROUP__', 'S型存储 (3条)'),
    ('sb',    '0100011', '000', '-',     'PC4', '0', '-',      'S_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', '-',    '0',       'WE_B'),
    ('sh',    '0100011', '001', '-',     'PC4', '0', '-',      'S_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', '-',    '0',       'WE_H'),
    ('sw',    '0100011', '010', '-',     'PC4', '0', '-',      'S_TYPE', 'ADD', 'SEL_RS1', 'SEL_EXT', '-',    '0',       'WE_W'),

    ('__GROUP__', 'B型分支 (4条)'),
    ('blt',   '1100011', '100', '-',     'PC4', '0', '-',      'B_TYPE', 'SLT', 'SEL_RS1', 'SEL_RS2', 'BRANCH','0',      '0'),
    ('bltu',  '1100011', '110', '-',     'PC4', '0', '-',      'B_TYPE', 'SLTU','SEL_RS1', 'SEL_RS2', 'BRANCH','0',      '0'),
    ('bge',   '1100011', '101', '-',     'PC4', '0', '-',      'B_TYPE', 'SLT', 'SEL_RS1', 'SEL_RS2', 'BRANCH','0',      '0'),
    ('bgeu',  '1100011', '111', '-',     'PC4', '0', '-',      'B_TYPE', 'SLTU','SEL_RS1', 'SEL_RS2', 'BRANCH','0',      '0'),

    ('__GROUP__', 'U型 (1条)'),
    ('auipc', '0010111', '-',   '-',     'PC4', '1', 'WB_ALU', 'U_TYPE', 'ADD', 'SEL_PC',  'SEL_EXT', '-',    '0',       '0'),
]

r = 5
for item in ctrl_data:
    if item[0] == '__GROUP__':
        apply_merge(ws2, r, 2, r, 15, item[1], bold_font, group_fill, left_align)
        r += 1
        continue

    inst = item[0]
    fill = example_fill if inst in DEMO_SET else blank_fill
    for i, v in enumerate(item):
        apply_style(ws2, r, 2+i, v, normal_font, fill, left_align if i == 0 else center_align)
    r += 1

for row in range(3, r+1):
    ws2.row_dimensions[row].height = 20

# ============================================================
# SHEET 3: 信号编码定义
# ============================================================
ws3 = wb.create_sheet("信号编码定义")
for c, w in {1:5, 2:14, 3:14, 4:10, 5:52}.items():
    ws3.column_dimensions[get_column_letter(c)].width = w

r = 2
apply_merge(ws3, r, 1, r, 5, '控制信号编码定义 (与 defines.vh 一致)', title_font, title_fill)

def write_section(ws, start_row, title, headers, rows_data):
    r = start_row
    apply_merge(ws, r, 1, r, 5, title, Font(name='Microsoft YaHei', size=10, bold=True), None)
    r += 1
    for j, h in enumerate(headers):
        apply_style(ws, r, 1+j, h, header_font, header_fill)
    r += 1
    for row_data in rows_data:
        for j, v in enumerate(row_data):
            apply_style(ws, r, 1+j, v, normal_font)
        r += 1
    return r + 1

r = 4
r = write_section(ws3, r, 'npc_op / NPC 操作选择', ['信号值', '建议宏名', '编码', '说明', '适用指令'],
    [('PC4',    'NPC_PC4',  "2'b00", 'NPC = PC + 4', '大部分指令'),
     ('BRANCH', 'NPC_BRA',  "2'b10", 'NPC = br ? PC+offset : PC+4', 'B型分支 (beq/bne/blt/bge/bltu/bgeu)'),
     ('JUMP',   'NPC_JMP',  "2'b11", 'NPC = PC + offset', 'JAL; 对JALR实际NPC=ALU.C, 在CPU外部处理')])

r = write_section(ws3, r, 'rf_wsel / 写回数据选择', ['信号值', '建议宏名', '编码', '说明', '适用指令'],
    [('WB_ALU', 'WB_ALU', "2'b00", '写回 ALU.C (运算结果)', 'R型/I型算逻/AUIPC'),
     ('WB_RAM', 'WB_RAM', "2'b01", '写回 MEXT.ext (内存读取)', 'Load (lb/lbu/lh/lhu/lw)'),
     ('WB_PC4', 'WB_PC4', "2'b10", '写回 PC+4 (返回地址)', 'JAL/JALR'),
     ('WB_EXT', 'WB_EXT', "2'b11", '写回 SEXT.ext (立即数)', 'LUI')])

r = write_section(ws3, r, 'sext_op / 立即数扩展格式', ['信号值', '建议宏名', '编码', '说明', '适用指令'],
    [('I_TYPE', 'EXT_I', "3'b000", 'I型: {{20{inst[31]}}, inst[31:20]}', 'addi/ori/slli/andi/xori/srli/srai/slti/sltiu/lw/lb/lbu/lh/lhu/jalr'),
     ('S_TYPE', 'EXT_S', "3'b001", 'S型: {{20{inst[31]}}, inst[31:25], inst[11:7]}', 'sb/sh/sw'),
     ('B_TYPE', 'EXT_B', "3'b010", 'B型: {{19{inst[31]}}, inst[31], inst[7], inst[30:25], inst[11:8], 1\'b0}', 'beq/bne/blt/bge/bltu/bgeu'),
     ('U_TYPE', 'EXT_U', "3'b011", 'U型: {inst[31:12], 12\'b0}', 'lui/auipc'),
     ('J_TYPE', 'EXT_J', "3'b100", 'J型: {{11{inst[31]}}, inst[31], inst[19:12], inst[20], inst[30:21], 1\'b0}', 'jal')])

r = write_section(ws3, r, 'alu_op / ALU 运算选择', ['信号值', '建议宏名', '编码', '说明', '适用指令'],
    [('ADD',  'ALU_ADD',  "5'h00", 'C = A + B', 'add/addi/auipc/lw/lb/lbu/lh/lhu/sb/sh/sw/jal/jalr'),
     ('SUB',  'ALU_SUB',  "5'h01", 'C = A - B', 'sub'),
     ('AND',  'ALU_AND',  "5'h02", 'C = A & B', 'and/andi'),
     ('OR',   'ALU_OR',   "5'h03", 'C = A | B', 'or/ori'),
     ('XOR',  'ALU_XOR',  "5'h04", 'C = A ^ B', 'xor/xori'),
     ('SLL',  'ALU_SLL',  "5'h05", 'C = A << B[4:0]', 'sll/slli'),
     ('SRL',  'ALU_SRL',  "5'h06", 'C = A >> B[4:0] (逻辑)', 'srl/srli'),
     ('SRA',  'ALU_SRA',  "5'h07", 'C = $signed(A) >>> B[4:0]', 'sra/srai'),
     ('EQ',   'ALU_EQ',   "5'h08", 'br = (A == B)', 'beq'),
     ('NE',   'ALU_NE',   "5'h09", 'br = (A != B)', 'bne'),
     ('SLT',  'ALU_SLT',  "5'h0A", 'br/C = 有符号比较', 'slt/slti/blt/bge'),
     ('SLTU', 'ALU_SLTU', "5'h0B", 'br/C = 无符号比较', 'sltu/sltiu/bltu/bgeu'),
     ('MUL',  'ALU_MUL',  "5'h0C", 'C = (A*B)[31:0]', 'mul'),
     ('MULH', 'ALU_MULH', "5'h0D", 'C = (A*B)[63:32] 有符号', 'mulh'),
     ('MULHU','ALU_MULHU',"5'h0E", 'C = (A*B)[63:32] 无符号', 'mulhu'),
     ('DIV',  'ALU_DIV',  "5'h0F", 'C = A÷B 有符号', 'div'),
     ('DIVU', 'ALU_DIVU', "5'h10", 'C = A÷B 无符号', 'divu'),
     ('REM',  'ALU_REM',  "5'h11", 'C = A%B 有符号', 'rem'),
     ('REMU', 'ALU_REMU', "5'h12", 'C = A%B 无符号', 'remu')])

r = write_section(ws3, r, 'alua_sel / ALU A端口选择', ['信号值', '建议宏名', '编码', '说明', '适用指令'],
    [('SEL_RS1', 'ALU_A_RS1', "1'b0", 'ALU.A = RF.rD1 (rs1值)', '大部分指令'),
     ('SEL_PC',  'ALU_A_PC',  "1'b1", 'ALU.A = PC (当前PC)', 'auipc')])

r = write_section(ws3, r, 'alub_sel / ALU B端口选择', ['信号值', '建议宏名', '编码', '说明', '适用指令'],
    [('SEL_RS2', 'ALU_B_RS2', "1'b0", 'ALU.B = RF.rD2 (rs2值)', 'R型/B型'),
     ('SEL_EXT', 'ALU_B_EXT', "1'b1", 'ALU.B = SEXT.ext (立即数)', 'I型/S型/U型/J型')])

r = write_section(ws3, r, 'ram_rop / 访存读操作类型', ['信号值', '建议宏名', '编码', '说明', '适用指令'],
    [('0',       'RAM_EXT_N',  "3'b000", '不读', '大部分指令'),
     ('WORD_EXT','RAM_EXT_W',  "3'b001", '读字 (32-bit)', 'lw'),
     ('BYTE_EXT','RAM_EXT_B',  "3'b010", '读字节, 有符号扩展', 'lb'),
     ('BYTEU_EXT','RAM_EXT_BU',"3'b011", '读字节, 无符号扩展', 'lbu'),
     ('HALF_EXT','RAM_EXT_H',  "3'b100", '读半字, 有符号扩展', 'lh'),
     ('HALFU_EXT','RAM_EXT_HU',"3'b101", '读半字, 无符号扩展', 'lhu')])

r = write_section(ws3, r, 'ram_wop / 访存写操作类型', ['信号值', '建议宏名', '编码', '说明', '适用指令'],
    [('0',    'RAM_WE_N', "4'b0000", '不写', '大部分指令'),
     ('WE_B', 'RAM_WE_B', "4'b0001", '写字节', 'sb'),
     ('WE_H', 'RAM_WE_H', "4'b0011", '写半字', 'sh'),
     ('WE_W', 'RAM_WE_W', "4'b1111", '写字', 'sw')])

r = write_section(ws3, r, '地址空间映射 (Address Space)', ['区域', '建议宏名', '基地址', '大小', '说明'],
    [('Block Memory','MEM_BLOCK_MEMORY',"32'h0000_0000",'512KB','指令/数据存储器'),
     ('DDR3',       'MEM_DDR3',        "32'h2000_0000",'512MB','DDR3 SDRAM'),
     ('Switch',     'PERI_ADDR_SWITCH',"32'hFFFF_0000",'-',    '拨码开关'),
     ('LED',        'PERI_ADDR_LED',   "32'hFFFF_1000",'-',    'LED'),
     ('DigLED',     'PERI_ADDR_DIGLED',"32'hFFFF_2000",'-',    '数码管'),
     ('UART',       'PERI_ADDR_UART',  "32'hFFFF_3000",'-',    'UART串口'),
     ('Timer',      'PERI_ADDR_TIMER', "32'hFFFF_4000",'-',    '定时器')])

# Save
filepath = 'datapath-and-control-tables.xlsx'
wb.save(filepath)
print(f'Saved: {filepath}')
print(f'Sheet 1 "数据通路表(含控制信号)": {ws1.max_row} rows')
print(f'Sheet 2 "控制信号取值表": {ws2.max_row} rows')
print(f'Sheet 3 "信号编码定义": {ws3.max_row} rows')
